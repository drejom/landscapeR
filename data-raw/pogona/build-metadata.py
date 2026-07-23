#!/usr/bin/env python3
"""Build analysis metadata for the three local Pogona matrices."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "source_metadata"


def matrix_samples(directory: Path) -> list[str]:
    def header(path: Path) -> list[str]:
        with path.open(encoding="utf-8") as handle:
            return handle.readline().rstrip("\n").split("\t")[2:]

    counts = header(directory / "gene_counts.tsv")
    tpm = header(directory / "gene_tpm.tsv")
    if counts != tpm:
        raise ValueError(f"count/TPM columns differ in {directory.name}")
    return counts


FIELDS = [
    "experiment", "matrix_sample_id", "include", "exclusion_reason",
    "genotype", "temperature_c", "stage", "day", "tissue",
    "biological_unit", "source_group", "source_deg_id", "metadata_status",
]


def write_metadata(directory: Path, rows: list[dict[str, object]]) -> None:
    ids = matrix_samples(directory)
    if [row["matrix_sample_id"] for row in rows] != ids:
        raise ValueError(f"metadata order does not match matrix in {directory.name}")
    with (directory / "sample_metadata.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def build_timecourse() -> None:
    directory = ROOT / "gsd_timecourse_28c"
    workbook = load_workbook(
        SOURCE / "sample_metadata.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["SAMPLE MATCH MASTER"]
    values = list(sheet.iter_rows(values_only=True))
    columns = values[0]
    source_rows = {
        row[0]: dict(zip(columns, row)) for row in values[1:] if row[0]
    }
    rows = []
    for sample_id in matrix_samples(directory):
        source = source_rows.get(sample_id)
        if source is None:
            raise ValueError(f"time-course sample lacks metadata: {sample_id}")
        group = str(source["GROUP"])
        deg_id = str(source["DEG_ID"])
        group_match = re.fullmatch(r"(ZZ|ZW)_(\d+)", group)
        deg_match = re.fullmatch(r"(ZZ|ZW)_(\d+)_Rep\d+", deg_id)
        if group_match is None or deg_match is None:
            raise ValueError(f"unparseable time-course labels for {sample_id}")
        conflict = group_match.groups() != deg_match.groups()
        rows.append({
            "experiment": "gsd_timecourse_28c",
            "matrix_sample_id": sample_id,
            "include": str(not conflict).lower(),
            "exclusion_reason": "group/DEG_ID conflict" if conflict else "",
            "genotype": group_match.group(1),
            "temperature_c": 28,
            "stage": "",
            "day": group_match.group(2),
            "tissue": "not supplied",
            "biological_unit": "independent destructive sample or declared pool",
            "source_group": group,
            "source_deg_id": deg_id,
            "metadata_status": "group_deg_conflict" if conflict else "matched",
        })
    write_metadata(directory, rows)


def build_early() -> None:
    directory = ROOT / "early_urogenital_28c"
    rows = []
    for sample_id in matrix_samples(directory):
        match = re.fullmatch(r".+_S([124])_(ZZ|ZW)", sample_id)
        if match is None:
            raise ValueError(f"unparseable early-stage sample: {sample_id}")
        stage, genotype = match.groups()
        rows.append({
            "experiment": "early_urogenital_28c",
            "matrix_sample_id": sample_id,
            "include": "true",
            "exclusion_reason": "",
            "genotype": genotype,
            "temperature_c": 28,
            "stage": f"S{stage}",
            "day": "",
            "tissue": "whole urogenital system",
            "biological_unit": "independent destructive sample",
            "source_group": f"{genotype}_28C_S{stage}",
            "source_deg_id": "",
            "metadata_status": "parsed_from_matrix_id",
        })
    write_metadata(directory, rows)


def build_gonad() -> None:
    directory = ROOT / "gonad_temperature"
    with (SOURCE / "sample_metadata_groups.csv").open(encoding="utf-8-sig") as handle:
        source_rows = {row["DEG_ID"]: row for row in csv.DictReader(handle)}
    rows = []
    for sample_id in matrix_samples(directory):
        source = source_rows.get(sample_id)
        if source is None:
            body = "body" in sample_id.lower()
            rows.append({
                "experiment": "gonad_temperature",
                "matrix_sample_id": sample_id,
                "include": "false",
                "exclusion_reason": "non-gonad body sample" if body else "no supplied metadata match",
                "genotype": "", "temperature_c": "", "stage": "", "day": "",
                "tissue": "body" if body else "unknown",
                "biological_unit": "independent destructive sample",
                "source_group": "", "source_deg_id": "",
                "metadata_status": "unmatched",
            })
            continue
        rows.append({
            "experiment": "gonad_temperature",
            "matrix_sample_id": sample_id,
            "include": "true",
            "exclusion_reason": "",
            "genotype": source["Genotype"],
            "temperature_c": source["TempGrp"].removesuffix("C"),
            "stage": source["Stage"],
            "day": "",
            "tissue": source["Tissue"],
            "biological_unit": "independent destructive sample",
            "source_group": source["GROUP"],
            "source_deg_id": source["DEG_ID"],
            "metadata_status": "matched",
        })
    write_metadata(directory, rows)


if __name__ == "__main__":
    build_timecourse()
    build_early()
    build_gonad()
